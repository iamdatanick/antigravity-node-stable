"""
Generate Antigravity Node v14.1 Phoenix Specification Workbook
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os

def auto_fit_columns(ws, min_width=12, max_width=50):
    """Auto-fit column widths based on content"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width

def style_sheet(ws, tab_color):
    """Apply consistent styling to a worksheet"""
    # Header styling
    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Alternating row colors
    light_blue_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        if row_idx % 2 == 0:
            for cell in row:
                cell.fill = light_blue_fill

    # Auto-filter
    ws.auto_filter.ref = ws.dimensions

    # Freeze panes
    ws.freeze_panes = "A2"

    # Tab color
    ws.sheet_properties.tabColor = tab_color

    # Auto-fit columns
    auto_fit_columns(ws)

def create_blockers_sheet(wb):
    """Sheet 1: Blockers"""
    ws = wb.create_sheet("Blockers", 0)

    # Headers
    headers = ["ID", "Title", "Severity", "Status", "Resolution / Rationale", "Target"]
    ws.append(headers)

    # Data
    data = [
        ["BLK-001", "OVMS Empty Config", "P0", "CLOSED", "scripts/init_models.sh generates model_config.json.", "MVP_Cloud"],
        ["BLK-002", "No Local GenAI", "P0", "CLOSED", "OVMS 2025.4 native GenAI support enabled.", "MVP_Cloud"],
        ["BLK-010", "A2A v0.3 Compliance", "P0", "CLOSED", "main.py refactored to use a2a-python SDK.", "MVP_Cloud"],
        ["BLK-016", "Multi-Tenant Storage", "P0", "CLOSED", "Ceph S3 Buckets created via setup_tenants.sh.", "MVP_Cloud"],
        ["BLK-017", "OpenBao Dev Mode", "P0", "CLOSED", "Forced Root Token -dev-root-token-id=root.", "MVP_Cloud"],
        ["BLK-023", "Etcd Security", "P0", "CLOSED", "Upgraded to v3.5.17.", "MVP_Cloud"],
        ["BLK-027", "MinIO AGPL Risk", "P0", "CLOSED", "Replaced with Ceph (LGPL/Apache).", "MVP_Cloud"],
        ["BLK-028", "Argo K8s Dependency", "P0", "CLOSED", "Replaced with AsyncDAGEngine (Python).", "MVP_Cloud"],
        ["BLK-029", "Missing Orchestrator", "P0", "CLOSED", "Implemented src/orchestrator/engine.py.", "MVP_Cloud"],
        ["BLK-030", "AVX-512 Crash Risk", "P0", "CLOSED", "check_avx512.sh guardrail added.", "MVP_Cloud"],
        ["BLK-031", "Ceph CORS Issue", "P0", "CLOSED", "s3cmd setcors added to entrypoint.", "MVP_Cloud"],
        ["BLK-034", "Phantom DB Dependency", "P0", "CLOSED", "Removed asyncpg; Switched to etcd3 for State.", "MVP_Cloud"],
        ["BLK-035", "Protocol Mismatch", "P0", "CLOSED", "Changed Proxy Target to HTTP/REST (Port 9001).", "MVP_Cloud"],
        ["BLK-036", "Data Persistence", "P0", "CLOSED", "Added Docker Volumes for Ceph (Data+Config) & Etcd.", "MVP_Cloud"],
        ["BLK-037", "Container Permissions", "P0", "CLOSED", "Added scripts/fix_perms.sh to pre-boot sequence.", "MVP_Cloud"],
        ["BLK-038", "Dependency Ambiguity", "P0", "CLOSED", "Pinned python-etcd3 & added smoke test.", "MVP_Cloud"],
    ]

    for row in data:
        ws.append(row)

    style_sheet(ws, "FF0000")  # Red

def create_actions_sheet(wb):
    """Sheet 2: Actions"""
    ws = wb.create_sheet("Actions")

    # Headers
    headers = ["ID", "Action", "Owner", "Deps", "Deadline", "Exit Criteria"]
    ws.append(headers)

    # Data
    data = [
        ["ACT-100", "Create scripts/check_avx512.sh", "Platform", "None", "T-0", "Script returns 0 on AVX-512 hosts; warns on others."],
        ["ACT-101", "Update requirements.txt", "Platform", "None", "T-0", "a2a-python>=0.3, python-etcd3, aioboto3 (Async S3)."],
        ["ACT-102", "Create src/orchestrator/engine.py", "Platform", "ACT-101", "T-0", "Engine uses etcd3 for locking; aioboto3 for S3."],
        ["ACT-103", "Refactor workflows/main.py", "Protocol", "ACT-101", "T-0", "Endpoints replaced with A2AFastAPI router."],
        ["ACT-104", "Update docker-compose.yml", "DevOps", "None", "T-0", "Map Ceph (Data+Conf) / Etcd Volumes; Set Vault Root Token; Healthchecks added; restart: unless-stopped."],
        ["ACT-105", "Create scripts/init_models.sh", "Platform", "ACT-104", "T-1", "TinyLlama GGUF + model_config.json generated."],
        ["ACT-106", "Create scripts/setup_tenants.sh", "Security", "ACT-104", "T-1", "User tenant-a created in Ceph RGW."],
        ["ACT-107", "Configure budget-proxy/routes.yaml", "Platform", "None", "T-1", "Route tinyllama to http://ovms:9001 (REST)."],
        ["ACT-108", "Implement MCP Bridge", "Integration", "ACT-102", "T-2", "engine.py calls MCP tools dynamically."],
        ["ACT-109", "Create scripts/fix_perms.sh", "DevOps", "ACT-104", "T-0", "Script chowns ./data/ directories to container UIDs."],
        ["ACT-110", "Update Deployment_Order logic", "DevOps", "ACT-105", "T-0", "setup_tenants.sh waits for ceph healthcheck. engine.py waits for etcd healthcheck."],
        ["ACT-111", "Verify Package Imports", "Platform", "ACT-101", "T-0", 'CI/Script runs python -c "import etcd3; ..." to catch naming ambiguity.'],
    ]

    for row in data:
        ws.append(row)

    style_sheet(ws, "FFA500")  # Orange

def create_code_gaps_sheet(wb):
    """Sheet 3: Code_Gaps"""
    ws = wb.create_sheet("Code_Gaps")

    # Headers
    headers = ["ID", "File Path", "Change Specification"]
    ws.append(headers)

    # Data
    data = [
        ["CG-101", "requirements.txt", 'Remove asyncpg. Add python-etcd3==0.12.0 (State), aioboto3==11.3.0 (S3), boto3, tenacity==8.2.3, grpcio==1.60.0 (Pinned). Smoke test: python -c "import etcd3; import aioboto3; import grpc".'],
        ["CG-102", "src/orchestrator/engine.py", "Use etcd3 client for Distributed Locking. Use aioboto3 for Artifact persistence."],
        ["CG-103", "workflows/main.py", 'Remove @app.post("/task"). Add app.include_router(A2AFastAPI(agent).router).'],
        ["CG-104", "docker-compose.yml", "Ceph: Vols ./data/ceph:/var/lib/ceph, ./data/ceph_conf:/etc/ceph. Etcd: Vol ./data/etcd:/etcd-data, Env ETCD_DATA_DIR=/etcd-data, Command /usr/local/bin/etcd --data-dir=/etcd-data. OpenBao: Command server -dev -dev-root-token-id=root."],
        ["CG-105", "docker-compose.yml", "OVMS: Expose Port 9001 (REST). Env ONEDNN_MAX_CPU_ISA=AVX512_CORE. CMD: --config_path /models/model_config.json. Healthchecks on etcd (etcdctl endpoint health), ceph (curl -f http://localhost:8000), ovms (curl -f http://localhost:9001/v2/health/live). restart: unless-stopped for all."],
        ["CG-106", "config/budget-proxy.yaml", "Add routing rule: model: tinyllama -> target: http://ovms:9001."],
        ["CG-107", "scripts/check_avx512.sh", "Enforce: If AVX-512 not found, exit 1 (Fail Fast). User must override with FORCE_NO_AVX=1."],
        ["CG-108", "scripts/fix_perms.sh", "Create script to chown -R 167:167 ./data/ceph and chown -R 1001:1001 ./data/etcd before boot."],
    ]

    for row in data:
        ws.append(row)

    style_sheet(ws, "0000FF")  # Blue

def create_deployment_order_sheet(wb):
    """Sheet 4: Deployment_Order"""
    ws = wb.create_sheet("Deployment_Order")

    # Headers
    headers = ["Phase", "Step", "Service/Script", "Command", "Validation"]
    ws.append(headers)

    # Data
    data = [
        [0, 1, "Hardware Check", "./scripts/check_avx512.sh", 'Output: "HARDWARE PASS" (or explicit override)'],
        [0, 2, "Permission Fix", "./scripts/fix_perms.sh", "./data owned by correct UIDs"],
        [0, 3, "Dependency Install", "pip install -r requirements.txt", 'python -c "import etcd3; import grpc" passes'],
        [0, 4, "Model Hydration", "./scripts/init_models.sh", "model_config.json & GGUF exist"],
        [1, 5, "Infrastructure Boot", "docker compose up -d ceph-demo etcd openbao", "Wait for docker compose ps -> healthy"],
        [1, 6, "Security Init", "./scripts/setup_tenants.sh", "s3cmd ls shows buckets (Retries until success)"],
        [2, 7, "Full Stack Boot", "docker compose up -d", "All containers up & healthy"],
        [3, 8, "Health Check", "curl localhost:8080/health", '{"status": "healthy"}'],
        [4, 9, "E2E Test", 'Chat UI -> "Why is sky blue?"', "Response from Local LLM via REST"],
    ]

    for row in data:
        ws.append(row)

    style_sheet(ws, "00FF00")  # Green

def create_sdk_requirements_sheet(wb):
    """Sheet 5: SDK_Requirements"""
    ws = wb.create_sheet("SDK_Requirements")

    # Headers
    headers = ["Package", "Version", "License", "Foundation", "Status"]
    ws.append(headers)

    # Data
    data = [
        ["python-etcd3", "0.12.0", "Apache-2.0", "Independent", "APPROVED"],
        ["aioboto3", "11.3.0", "Apache-2.0", "Independent", "APPROVED"],
        ["a2a-python", "0.3.0", "Apache-2.0", "Linux Foundation", "APPROVED"],
        ["openvino-genai", "2025.4.0", "Apache-2.0", "Intel OSS", "APPROVED"],
        ["optimum-intel", "1.22.0", "Apache-2.0", "Intel OSS", "APPROVED"],
        ["fastapi", "0.115.6", "MIT", "Independent", "APPROVED"],
        ["grpcio", "1.60.0", "Apache-2.0", "CNCF", "APPROVED"],
        ["pydantic", "2.10.5", "MIT", "Independent", "APPROVED"],
        ["tenacity", "8.2.3", "Apache-2.0", "Independent", "APPROVED"],
        ["boto3", "1.35.0", "Apache-2.0", "Independent", "APPROVED"],
        ["asyncpg", "REMOVED", "N/A", "N/A", "REMOVED"],
    ]

    for row in data:
        ws.append(row)

    style_sheet(ws, "800080")  # Purple

def create_architecture_gaps_sheet(wb):
    """Sheet 6: Architecture_Gaps"""
    ws = wb.create_sheet("Architecture_Gaps")

    # Headers
    headers = ["ID", "Gap", "Solution in v14.1 (Cloud)", "Solution in v15.0 (Bare Metal)"]
    ws.append(headers)

    # Data
    data = [
        ["GAP-012", "Storage Isolation", "Logical: Ceph S3 Policy + Tenant Buckets", "Physical: Rook-Ceph Block Devices"],
        ["GAP-025", "Orchestration", "Lightweight: AsyncDAGEngine (Etcd State)", "Heavy: Argo Workflows (K8s)"],
        ["GAP-030", "Protocol Mismatch", "Adapter: Budget Proxy (REST -> REST)", "Native: Mesh gRPC (Envoy/Cilium)"],
        ["GAP-031", "Secret bootstrapping", "Static: Root Token in Compose", "Dynamic: Vault Agent Injector"],
    ]

    for row in data:
        ws.append(row)

    style_sheet(ws, "FFFF00")  # Yellow

def create_validation_gates_sheet(wb):
    """Sheet 7: Validation_Gates"""
    ws = wb.create_sheet("Validation_Gates")

    # Headers
    headers = ["ID", "Gate Name", "Method", "Pass Criteria"]
    ws.append(headers)

    # Data
    data = [
        ["VG-101", "Hardware Compatibility", "Run check_avx512.sh", "Exit Code 0"],
        ["VG-102", "Protocol Compliance", "curl /tasks/send", "JSON-RPC 2.0 Response (ID; Result)"],
        ["VG-103", "Local Intelligence", "Chat UI Prompt", "Response generated without WAN traffic"],
        ["VG-104", "Data Sovereignty", "Check Ceph Bucket", "JSON Artifact found in tenant-a/"],
        ["VG-105", "Security", "Check OpenBao", "Secrets retrieved via Agent"],
        ["VG-106", "Resilience (Restart)", "docker compose restart", "Etcd Lock holds; Ceph Data/Keys persist; No Permission Denied errors."],
        ["VG-107", "Resilience (Recreate)", "docker compose down && docker compose up -d", "Data persists (Volume check); Identity stable."],
        ["VG-108", "Dependency Smoke", 'python -c "import ..."', "Exit Code 0 (No ImportError)."],
        ["VG-109", "AVX-512 Guard", "check_avx512.sh", "Exits 1 on non-compatible hardware (Fail Safe)."],
    ]

    for row in data:
        ws.append(row)

    style_sheet(ws, "008080")  # Teal

def main():
    """Main function to create the workbook"""
    output_path = r"C:\Users\NickV\OneDrive\Desktop\Antigravity-Node\docs\AntigravityNode_v14.1_Phoenix_Spec.xlsx"

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    # Create all sheets
    create_blockers_sheet(wb)
    create_actions_sheet(wb)
    create_code_gaps_sheet(wb)
    create_deployment_order_sheet(wb)
    create_sdk_requirements_sheet(wb)
    create_architecture_gaps_sheet(wb)
    create_validation_gates_sheet(wb)

    # Save workbook
    wb.save(output_path)

    # Get file size
    file_size = os.path.getsize(output_path)

    print("SUCCESS: Workbook created successfully!")
    print(f"File path: {output_path}")
    print(f"File size: {file_size:,} bytes ({file_size / 1024:.2f} KB)")
    print(f"Sheets created: {len(wb.sheetnames)}")
    print(f"Sheet names: {', '.join(wb.sheetnames)}")

if __name__ == "__main__":
    main()
